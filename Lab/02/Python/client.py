from openpyxl import Workbook
from openpyxl import load_workbook
from flask import Flask, request
import json
import datetime


app = Flask(__name__)
jbuffer = []
@app.route('/', methods=['POST'])
def index():
	rows = request.json
	status = "Ok"
	now = datetime.datetime.now()#текущее время
	for row in rows['check']: # получение новых данных
		dicti = {"user_id":rows['user_id'],"datetime":now.strftime("%d.%m.%Y %H:%M:%S"),"item":row['item'],"price":row['price']}
		jbuffer.append(dicti)
		if len(jbuffer) >= 10:
			try: # проверка на наличие файла
				book = load_workbook("data.xlsx")
			except FileNotFoundError:
				#Если файла нет
				wb = Workbook()
				ws = wb.active
				ws['A1'] = 'N';
				ws['B1'] = 'User ID';
				ws['C1'] = 'Datetime';
				ws['D1'] = 'Item';
				ws['E1'] = 'Prise';
				wb.save("data.xlsx")
			else:
				book.close()
			book = load_workbook("data.xlsx")
			sheet = book.active#достаем данные
			ink = sheet.max_row
			for row in jbuffer:
				sheet[1+ink][0].value = ink;
				sheet[1+ink][1].value = row['user_id'];
				sheet[1+ink][2].value = row['datetime']
				sheet[1+ink][3].value = row['item'];
				sheet[1+ink][4].value = row['price'];
				ink +=1
			try: # проверка на наличие файла
				book.save("data.xlsx")
			except PermissionError:
				status = "permission error"
				break
			else:
				jbuffer.clear()
				status = "added"
	return status

if __name__ == "__main__":
    app.run()
